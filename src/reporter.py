# src/reporter.py
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_summary_sheet(workbook, analysis_results: Dict, conversation_stats: Dict):
    """Create a summary overview sheet with key metrics."""

    ws = workbook.create_sheet("📊 Summary", 0)

    # Title
    ws['A1'] = 'LinkedIn DM Analysis Summary'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    # Generation timestamp
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True)

    # Key Metrics Section
    row = 4
    ws[f'A{row}'] = 'KEY METRICS'
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    metrics = [
        ('Total Messages', analysis_results.get('total_messages', 0)),
        ('Total Contacts', conversation_stats.get('total_contacts', 0)),
        ('Outbound Messages', analysis_results.get('outbound_message_count', 0)),
        ('Inbound Messages', analysis_results.get('inbound_message_count', 0)),
        ('', ''),  # Spacer
        ('Overall Response Rate', f"{conversation_stats.get('overall_response_rate', 0):.1%}"),
        ('Contacts Who Responded', conversation_stats.get('contacts_who_responded', 0)),
        ('Contacts Who Ghosted', conversation_stats.get('contacts_who_ghosted', 0)),
        ('', ''),  # Spacer
        ('Avg Response Time (hours)', f"{analysis_results.get('avg_response_time_hours', 0):.1f}"),
        ('Quick Responses (<1h)', f"{analysis_results.get('quick_response_rate', 0):.1%}"),
    ]

    for metric, value in metrics:
        if metric:  # Skip spacers
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            if isinstance(value, str) and '%' in value:
                ws[f'B{row}'].font = Font(bold=True)
        row += 1

    # Message Quality Section
    row += 1
    ws[f'A{row}'] = 'MESSAGE QUALITY'
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    quality_metrics = [
        ('Avg Outbound Length (chars)', f"{analysis_results.get('outbound_avg_length', 0):.0f}"),
        ('Avg Inbound Length (chars)', f"{analysis_results.get('inbound_avg_length', 0):.0f}"),
        ('Avg Outbound Sentiment', f"{analysis_results.get('outbound_avg_sentiment', 0):.2f}"),
        ('Avg Inbound Sentiment', f"{analysis_results.get('inbound_avg_sentiment', 0):.2f}"),
    ]

    for metric, value in quality_metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        row += 1

    # Apply formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top')

def create_contact_analysis_sheet(workbook, contact_summary: pd.DataFrame):
    """Create detailed contact analysis sheet."""

    ws = workbook.create_sheet("👥 Contact Analysis")

    # Add title
    ws['A1'] = 'Contact Analysis'
    ws['A1'].font = Font(size=14, bold=True)

    # Prepare data for display
    display_df = contact_summary.copy()
    display_df['response_rate'] = display_df['response_rate'].apply(lambda x: f"{x:.1%}")
    display_df['first_contact'] = display_df['first_contact'].dt.strftime('%Y-%m-%d')
    display_df['last_contact'] = display_df['last_contact'].dt.strftime('%Y-%m-%d')

    # Column headers mapping
    column_mapping = {
        'contact_person': 'Contact Name',
        'total_messages': 'Total Messages',
        'outbound_messages': 'Sent by You',
        'inbound_messages': 'Received',
        'response_rate': 'Response Rate',
        'has_response': 'Responded?',
        'thread_count': 'Conversations',
        'first_contact': 'First Contact',
        'last_contact': 'Last Contact'
    }

    display_df = display_df.rename(columns=column_mapping)

    # Add data to sheet
    for r in dataframe_to_rows(display_df, index=False, header=True):
        ws.append(r)

    # Format headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[2]:  # Header row
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_top_messages_sheet(workbook, top_messages: pd.DataFrame):
    """Create sheet with top performing messages."""

    ws = workbook.create_sheet("🎯 Top Messages")

    if len(top_messages) == 0:
        ws['A1'] = 'No top performing messages found'
        return

    # Add title
    ws['A1'] = 'Top Performing Messages'
    ws['A1'].font = Font(size=14, bold=True)

    # Prepare data
    display_df = top_messages.copy()
    display_df['timestamp'] = display_df['timestamp'].dt.strftime('%Y-%m-%d %H:%M')
    display_df['performance_score'] = display_df['performance_score'].round(2)
    display_df['sentiment_polarity'] = display_df['sentiment_polarity'].round(3)

    # Limit content length for display
    display_df['content'] = display_df['content'].apply(
        lambda x: x[:200] + '...' if len(str(x)) > 200 else x
    )

    column_mapping = {
        'content': 'Message Content',
        'contact_person': 'Contact',
        'performance_score': 'Score',
        'sentiment_polarity': 'Sentiment',
        'message_length': 'Length',
        'word_count': 'Words',
        'timestamp': 'Sent Date',
        'has_greeting': 'Has Greeting',
        'question_count': 'Questions'
    }

    display_df = display_df.rename(columns=column_mapping)

    # Add data
    for r in dataframe_to_rows(display_df, index=False, header=True):
        ws.append(r)

    # Format headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    column_widths = {
        'A': 60,  # Message content
        'B': 25,  # Contact
        'C': 10,  # Score
        'D': 12,  # Sentiment
        'E': 10,  # Length
        'F': 10,  # Words
        'G': 18,  # Date
        'H': 12,  # Greeting
        'I': 12   # Questions
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Wrap text for message content
    for row in ws.iter_rows(min_row=3, max_col=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

def create_detailed_messages_sheet(workbook, analyzed_df: pd.DataFrame):
    """Create sheet with all analyzed messages."""

    ws = workbook.create_sheet("📋 All Messages")

    # Add title
    ws['A1'] = 'Detailed Message Analysis'
    ws['A1'].font = Font(size=14, bold=True)

    # Select and prepare columns
    columns_to_include = [
        'timestamp', 'thread_id', 'contact_person', 'direction',
        'content', 'message_length', 'word_count',
        'sentiment_label', 'sentiment_polarity',
        'is_first_message', 'is_reply', 'response_time_hours',
        'has_greeting', 'question_count', 'has_url'
    ]

    display_df = analyzed_df[columns_to_include].copy()

    # Format timestamp
    display_df['timestamp'] = display_df['timestamp'].dt.strftime('%Y-%m-%d %H:%M')

    # Limit content length
    display_df['content'] = display_df['content'].apply(
        lambda x: str(x)[:100] + '...' if len(str(x)) > 100 else str(x)
    )

    # Round numeric columns
    numeric_columns = ['sentiment_polarity', 'response_time_hours']
    for col in numeric_columns:
        if col in display_df.columns:
            # Convert to numeric first, then round
            display_df[col] = pd.to_numeric(display_df[col], errors='coerce').round(2)

    column_mapping = {
        'timestamp': 'Date/Time',
        'thread_id': 'Thread ID',
        'contact_person': 'Contact',
        'direction': 'Direction',
        'content': 'Message (Preview)',
        'message_length': 'Length',
        'word_count': 'Words',
        'sentiment_label': 'Sentiment',
        'sentiment_polarity': 'Sentiment Score',
        'is_first_message': 'First Message?',
        'is_reply': 'Reply?',
        'response_time_hours': 'Response Time (h)',
        'has_greeting': 'Has Greeting?',
        'question_count': 'Questions',
        'has_url': 'Has URL?'
    }

    display_df = display_df.rename(columns=column_mapping)

    # Add data
    for r in dataframe_to_rows(display_df, index=False, header=True):
        ws.append(r)

    # Format headers
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths with limits
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

def generate_excel_report(analyzed_df: pd.DataFrame, contact_summary: pd.DataFrame,
                         analysis_results: Dict, conversation_stats: Dict,
                         top_messages: pd.DataFrame, output_path: str):
    """
    Generate comprehensive Excel report with multiple sheets.

    Args:
        analyzed_df: DataFrame with all analyzed messages
        contact_summary: Summary statistics by contact
        analysis_results: Dictionary with analysis results
        conversation_stats: Dictionary with conversation statistics
        top_messages: DataFrame with top performing messages
        output_path: Path for output Excel file
    """
    print(f"📊 Generating Excel report: {output_path}")

    # Create workbook
    workbook = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # Create all sheets
    create_summary_sheet(workbook, analysis_results, conversation_stats)
    create_contact_analysis_sheet(workbook, contact_summary)
    create_top_messages_sheet(workbook, top_messages)
    create_detailed_messages_sheet(workbook, analyzed_df)

    # Save workbook
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"✅ Excel report saved to: {output_path}")
    print(f"📋 Report includes {len(workbook.sheetnames)} sheets:")
    for sheet_name in workbook.sheetnames:
        print(f"   - {sheet_name}")

def print_quick_summary(analysis_results: Dict, conversation_stats: Dict):
    """Print a quick text summary to console."""

    print("\n" + "="*60)
    print("🎉 LINKEDIN DM ANALYSIS COMPLETE")
    print("="*60)

    print(f"\n📊 OVERVIEW:")
    print(f"   Total Messages: {analysis_results.get('total_messages', 0):,}")
    print(f"   Total Contacts: {conversation_stats.get('total_contacts', 0):,}")
    print(f"   Messages Sent: {analysis_results.get('outbound_message_count', 0):,}")
    print(f"   Messages Received: {analysis_results.get('inbound_message_count', 0):,}")

    print(f"\n💬 RESPONSE METRICS:")
    response_rate = conversation_stats.get('overall_response_rate', 0)
    print(f"   Overall Response Rate: {response_rate:.1%}")
    print(f"   Contacts Who Responded: {conversation_stats.get('contacts_who_responded', 0)}")
    print(f"   Contacts Who Ghosted: {conversation_stats.get('contacts_who_ghosted', 0)}")

    print(f"\n⏱️  TIMING:")
    avg_response = analysis_results.get('avg_response_time_hours', 0)
    print(f"   Avg Response Time: {avg_response:.1f} hours")
    quick_rate = analysis_results.get('quick_response_rate', 0)
    print(f"   Quick Response Rate (<1h): {quick_rate:.1%}")

    print(f"\n📝 MESSAGE QUALITY:")
    out_sentiment = analysis_results.get('outbound_avg_sentiment', 0)
    in_sentiment = analysis_results.get('inbound_avg_sentiment', 0)
    print(f"   Your Avg Sentiment: {out_sentiment:.2f}")
    print(f"   Contacts Avg Sentiment: {in_sentiment:.2f}")

    print("\n" + "="*60)